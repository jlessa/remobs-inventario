from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INVENTORY = ROOT / "docs" / "INVENTARIO.xlsx"
DEFAULT_FIELD = ROOT / "docs" / "Ficha_Campo_MEq_AX39_DD_MM_AA.xlsx"


def clean(value: Any) -> Any:
    if isinstance(value, ArrayFormula):
        return None
    if value is None:
        return None
    if isinstance(value, str):
        normalized = re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()
        return normalized or None
    return value


def text(value: Any) -> str | None:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def number(value: Any) -> int | None:
    value = clean(value)
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return max(0, int(value))
    if isinstance(value, str):
        candidate = value.strip().replace(",", ".")
        if candidate.lower() in {"n/a", "na", "nan", "-"}:
            return None
        try:
            return max(0, int(float(candidate)))
        except ValueError:
            return None
    return None


def marker_for(*parts: Any) -> str:
    raw = "|".join(text(part) or "" for part in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def status_from_condition(value: Any) -> str:
    normalized = (text(value) or "").casefold()
    if "avari" in normalized or "queb" in normalized:
        return "avariado"
    if "manuten" in normalized:
        return "manutencao"
    if "inconsist" in normalized:
        return "inconsistencia"
    return "operacional"


def sensor_status_from_condition(value: Any) -> str:
    normalized = (text(value) or "").casefold()
    if "avari" in normalized or "queb" in normalized:
        return "avariado"
    if "manuten" in normalized:
        return "manutencao"
    if "inconsist" in normalized:
        return "inconsistencia"
    return "nao_instalado"


def platform_status(value: Any) -> str:
    normalized = (text(value) or "").casefold()
    if "operacional" in normalized:
        return "em_operacao"
    if "desativ" in normalized or "nao existe" in normalized or "năo existe" in normalized:
        return "offline"
    if "planejamento" in normalized:
        return "planejamento"
    return "indefinido"


def compact_unit(value: Any, fallback: str = "un") -> str:
    unit = text(value) or fallback
    return unit[:32]


def append_description(item: dict[str, Any], line: str) -> None:
    current = item.setdefault("description_lines", [])
    if line and line not in current:
        current.append(line)


def merge_inventory_item(items: OrderedDict[str, dict[str, Any]], item: dict[str, Any]) -> None:
    source_key = item["source_key"]
    if source_key not in items:
        items[source_key] = item
        return

    existing = items[source_key]
    existing["source_rows"] = sorted(set((existing.get("source_rows") or []) + (item.get("source_rows") or [])))
    existing["quantity"] = max(int(existing.get("quantity") or 0), int(item.get("quantity") or 0))
    for field in ["condition_status", "brand", "model", "serial_number", "patrimony_number", "unit"]:
        if not existing.get(field) and item.get(field):
            existing[field] = item[field]
    for line in item.get("description_lines", []):
        append_description(existing, line)


def inventory_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    items: OrderedDict[str, dict[str, Any]] = OrderedDict()

    lab = workbook["INVENTARIO LAB"]
    for row_index in range(2, lab.max_row + 1):
        name = text(lab.cell(row_index, 1).value)
        category = text(lab.cell(row_index, 2).value)
        if not name or not category:
            continue
        brand = text(lab.cell(row_index, 3).value)
        model = text(lab.cell(row_index, 4).value)
        serial = text(lab.cell(row_index, 5).value)
        if serial and serial.casefold() in {"n/a", "na", "nan", "-"}:
            serial = None
        condition = text(lab.cell(row_index, 6).value)
        quantity = number(lab.cell(row_index, 7).value)
        test_date = text(lab.cell(row_index, 8).value)
        observation = text(lab.cell(row_index, 9).value)
        departure = text(lab.cell(row_index, 10).value)
        destination = text(lab.cell(row_index, 11).value)
        requester = text(lab.cell(row_index, 12).value)
        source_key = marker_for("inventory-lab", name, category, brand, model, serial or row_index)
        item = {
            "source_key": source_key,
            "source_file": path.name,
            "source_sheet": lab.title,
            "source_rows": [row_index],
            "item_type": "permanent_component" if category.casefold() != "consumo" else "consumable",
            "category_name": category,
            "location_name": "Laboratorio",
            "name": name,
            "brand": brand,
            "model": model,
            "serial_number": serial,
            "patrimony_number": None,
            "invoice_number": None,
            "condition_status": status_from_condition(condition),
            "unit": "un",
            "quantity": quantity if quantity is not None else 1,
            "minimum_stock_national": 0,
            "minimum_stock_import": 0,
            "minimum_stock_maintenance": 0,
            "ideal_stock": quantity if quantity is not None else 1,
            "description_lines": [
                f"Origem: {path.name} > {lab.title} linha {row_index}.",
                f"Condicao na planilha: {condition or 'nao informada'}.",
            ],
        }
        for label, value in [
            ("Data de teste", test_date),
            ("Observacao", observation),
            ("Saida", departure),
            ("Destino", destination),
            ("Solicitante", requester),
        ]:
            if value:
                append_description(item, f"{label}: {value}.")
        merge_inventory_item(items, item)

    paiol = workbook["INVENTARIO PAIOL"]
    for row_index in range(2, paiol.max_row + 1):
        name = text(paiol.cell(row_index, 1).value)
        category = text(paiol.cell(row_index, 2).value)
        if not name or not category:
            continue
        brand = text(paiol.cell(row_index, 3).value)
        model = text(paiol.cell(row_index, 4).value)
        dimension = text(paiol.cell(row_index, 5).value)
        unit = compact_unit(paiol.cell(row_index, 6).value)
        quantity = number(paiol.cell(row_index, 7).value)
        updated_at = text(paiol.cell(row_index, 8).value)
        source_key = marker_for("inventory-paiol", name, category, brand, model, dimension, unit)
        item = {
            "source_key": source_key,
            "source_file": path.name,
            "source_sheet": paiol.title,
            "source_rows": [row_index],
            "item_type": "permanent_component" if category.casefold() == "casco naval" else "consumable",
            "category_name": category,
            "location_name": "Paiol",
            "name": name,
            "brand": brand,
            "model": model,
            "serial_number": None,
            "patrimony_number": None,
            "invoice_number": None,
            "condition_status": "operacional",
            "unit": unit,
            "quantity": quantity or 0,
            "minimum_stock_national": 0,
            "minimum_stock_import": 0,
            "minimum_stock_maintenance": 0,
            "ideal_stock": quantity or 0,
            "description_lines": [
                f"Origem: {path.name} > {paiol.title} linha {row_index}.",
            ],
        }
        if dimension:
            append_description(item, f"Dimensao: {dimension}.")
        if updated_at:
            append_description(item, f"Ultima atualizacao: {updated_at}.")
        merge_inventory_item(items, item)

    page2 = workbook[workbook.sheetnames[2]]
    for row_index in range(2, page2.max_row + 1):
        name = text(page2.cell(row_index, 1).value)
        category = text(page2.cell(row_index, 2).value)
        if not name or not category:
            continue
        brand = text(page2.cell(row_index, 3).value)
        model = text(page2.cell(row_index, 4).value)
        condition = text(page2.cell(row_index, 6).value)
        bought = number(page2.cell(row_index, 7).value) or 0
        consumed = number(page2.cell(row_index, 8).value) or 0
        quantity = max(0, bought - consumed)
        source_key = marker_for("inventory-page2", name, category, brand, model)
        item = {
            "source_key": source_key,
            "source_file": path.name,
            "source_sheet": page2.title,
            "source_rows": [row_index],
            "item_type": "consumable",
            "category_name": category,
            "location_name": "Estoque",
            "name": name,
            "brand": brand,
            "model": model,
            "serial_number": None,
            "patrimony_number": None,
            "invoice_number": None,
            "condition_status": status_from_condition(condition),
            "unit": "un",
            "quantity": quantity,
            "minimum_stock_national": 0,
            "minimum_stock_import": 0,
            "minimum_stock_maintenance": 0,
            "ideal_stock": bought,
            "description_lines": [
                f"Origem: {path.name} > {page2.title} linha {row_index}.",
                f"Quantidade comprada: {bought}. Consumidos: {consumed}.",
            ],
        }
        merge_inventory_item(items, item)

    return [finalize_inventory_item(item) for item in items.values()]


def tool_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    if "Mala de ferramentas" not in workbook.sheetnames:
        return []
    worksheet = workbook["Mala de ferramentas"]
    items: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row_index in range(2, worksheet.max_row + 1):
        name = text(worksheet.cell(row_index, 1).value)
        if not name:
            continue
        tool_type = text(worksheet.cell(row_index, 2).value)
        model = text(worksheet.cell(row_index, 3).value)
        condition = text(worksheet.cell(row_index, 4).value)
        serial = text(worksheet.cell(row_index, 5).value)
        if serial and serial.casefold() in {"n/a", "na", "nan", "-"}:
            serial = None
        quantity = number(worksheet.cell(row_index, 6).value) or 1
        owner = text(worksheet.cell(row_index, 7).value)
        departure = text(worksheet.cell(row_index, 8).value)
        destination = text(worksheet.cell(row_index, 9).value)
        source_key = marker_for("toolbox", name, tool_type, model, serial or row_index)
        item = {
            "source_key": source_key,
            "source_file": path.name,
            "source_sheet": worksheet.title,
            "source_rows": [row_index],
            "item_type": "permanent_component",
            "category_name": "Ferramenta",
            "location_name": "Mala de ferramentas",
            "name": name,
            "brand": None,
            "model": " - ".join(part for part in [tool_type, model] if part) or None,
            "serial_number": serial,
            "patrimony_number": None,
            "invoice_number": None,
            "condition_status": status_from_condition(condition),
            "unit": "un",
            "quantity": quantity,
            "minimum_stock_national": 0,
            "minimum_stock_import": 0,
            "minimum_stock_maintenance": 0,
            "ideal_stock": quantity,
            "description_lines": [
                f"Origem: {path.name} > {worksheet.title} linha {row_index}.",
                f"Status na planilha: {condition or 'nao informado'}.",
            ],
        }
        if owner:
            append_description(item, f"Responsavel: {owner}.")
        if departure:
            append_description(item, f"Saida: {departure}.")
        if destination:
            append_description(item, f"Destino: {destination}.")
        merge_inventory_item(items, item)
    return [finalize_inventory_item(item) for item in items.values()]


def finalize_inventory_item(item: dict[str, Any]) -> dict[str, Any]:
    source_rows = sorted(set(item.get("source_rows") or []))
    if source_rows:
        item["source_rows"] = source_rows
    marker = f"remobs-import:{item['source_key']}"
    lines = item.pop("description_lines", [])
    lines.append(f"Marcador de importacao: {marker}.")
    item["description"] = "\n".join(lines)
    return item


def sensor_rows(inventory_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sensors: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for item in inventory_items:
        if item["source_sheet"] != "INVENTARIO LAB":
            continue
        if item["category_name"].casefold() != "sensor":
            continue
        source_key = marker_for("sensor", item["name"], item.get("brand"), item.get("model"), item.get("serial_number") or item["source_key"])
        note_lines = [
            f"Origem: {item['source_file']} > {item['source_sheet']} linhas {', '.join(map(str, item.get('source_rows') or []))}.",
            f"Item de inventario importado: {item['name']}.",
            f"Marcador de importacao: remobs-import:{source_key}.",
        ]
        if item.get("description"):
            note_lines.append(item["description"])
        sensors[source_key] = {
            "source_key": source_key,
            "sensor_type": item["name"],
            "family": item["name"],
            "brand": item.get("brand"),
            "model": item.get("model"),
            "serial_number": item.get("serial_number"),
            "patrimony_number": item.get("patrimony_number"),
            "operational_status": sensor_status_from_condition(item.get("condition_status")),
            "calibration_due_at": None,
            "notes": "\n".join(note_lines),
        }
    return list(sensors.values())


def platform_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook["Cadastro Estacoes"]
    platforms = []
    for row_index in range(2, worksheet.max_row + 1):
        station = text(worksheet.cell(row_index, 1).value)
        if not station:
            continue
        project = text(worksheet.cell(row_index, 2).value)
        client = text(worksheet.cell(row_index, 3).value)
        state = text(worksheet.cell(row_index, 4).value)
        region = text(worksheet.cell(row_index, 5).value)
        report_day = text(worksheet.cell(row_index, 6).value)
        lat = text(worksheet.cell(row_index, 7).value)
        lon = text(worksheet.cell(row_index, 8).value)
        status = text(worksheet.cell(row_index, 9).value)
        platform_type = text(worksheet.cell(row_index, 10).value) or "Estacao"
        product_type = text(worksheet.cell(row_index, 11).value)
        jira_label = text(worksheet.cell(row_index, 12).value)
        source_key = marker_for("platform", station)
        description_lines = [
            f"Origem: {path.name} > {worksheet.title} linha {row_index}.",
            f"Marcador de importacao: remobs-import:{source_key}.",
        ]
        for label, value in [
            ("Projeto", project),
            ("Cliente", client),
            ("UF", state),
            ("Regiao", region),
            ("Dia de emissao do relatorio", report_day),
            ("Latitude", lat),
            ("Longitude", lon),
            ("Status na planilha", status),
            ("Tipo de produto", product_type),
            ("Sigla Jira", jira_label),
        ]:
            if value:
                description_lines.append(f"{label}: {value}.")
        platforms.append(
            {
                "source_key": source_key,
                "name": station,
                "platform_type": platform_type,
                "manufacturer": client,
                "model": product_type,
                "operational_status": platform_status(status),
                "description": "\n".join(description_lines),
            }
        )
    return platforms


def checklist_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    rows = []
    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith("FichaCampo") or "TESTE" in sheet_name:
            continue
        worksheet = workbook[sheet_name]
        title = text(worksheet["A1"].value) or "Ficha de Campo"
        station = text(worksheet["F1"].value)
        if not station:
            continue
        trip = text(worksheet["B2"].value)
        team = text(worksheet["F2"].value)
        field_date = text(worksheet["I2"].value)
        nonempty_rows = []
        for row_index in range(1, worksheet.max_row + 1):
            values = [text(worksheet.cell(row_index, column).value) for column in range(1, min(worksheet.max_column, 14) + 1)]
            compact = [value for value in values if value and not value.startswith("#")]
            if compact:
                nonempty_rows.append({"linha": row_index, "valores": compact[:8]})
        source_key = marker_for("checklist", sheet_name, station)
        submitted = bool(field_date or team or trip)
        answers = {
            "operacao.codigo_viagem": trip,
            "operacao.equipe_hm": team,
            "operacao.data": field_date,
            "planilha.aba": sheet_name,
            "planilha.arquivo": path.name,
            "planilha.linhas_preenchidas": len(nonempty_rows),
            "estacao.codigo": station,
        }
        answers = {key: value for key, value in answers.items() if value not in (None, "")}
        if nonempty_rows:
            preview = []
            for row in nonempty_rows[:20]:
                preview.append(f"Linha {row['linha']}: {' | '.join(row['valores'])}")
            answers["planilha.resumo_linhas"] = "\n".join(preview)
        rows.append(
            {
                "source_key": source_key,
                "title": f"{title} - {station}",
                "template_name": title,
                "platform_name": station,
                "status": "submitted" if submitted else "draft",
                "current_step": 4 if submitted else 1,
                "total_steps": 4,
                "answers": answers,
                "evidence": [],
                "submitted_by_id": 0,
                "submitted_by_username": "importacao-planilhas",
                "notes": "\n".join(
                    [
                        f"Origem: {path.name} > {sheet_name}.",
                        f"Marcador de importacao: remobs-import:{source_key}.",
                        "Checklist importado da planilha de campo.",
                    ]
                ),
                "submitted_at": field_date if submitted and field_date else None,
            }
        )
    return rows


def build_payload(inventory_path: Path, field_path: Path) -> dict[str, Any]:
    inventory = inventory_rows(inventory_path)
    tools = tool_rows(field_path)
    all_items = inventory + tools
    return {
        "metadata": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "source_files": [str(inventory_path), str(field_path)],
        },
        "inventory_items": all_items,
        "sensors": sensor_rows(all_items),
        "platforms": platform_rows(field_path),
        "checklists": checklist_rows(field_path),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera payload JSON para importar planilhas REMOBS.")
    parser.add_argument("--inventory", type=Path, default=DEFAULT_INVENTORY)
    parser.add_argument("--field", type=Path, default=DEFAULT_FIELD)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    payload = build_payload(args.inventory, args.field)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    summary = {key: len(payload[key]) for key in ["inventory_items", "sensors", "platforms", "checklists"]}
    print(json.dumps(summary, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
